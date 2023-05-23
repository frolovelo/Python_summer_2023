import openpyxl
wb = openpyxl.load_workbook('zp.xlsx')
ws = wb.active
d = {}
for bk in wb.worksheets:
    wb.active = wb[bk.title]
    ws = wb.active
    for i in range(1, ws.max_row + 1):
        name = ws.cell(i, 1).value
        val = int(ws.cell(i, 2).value)
        d[name] = d.get(name, 0) + val

lst = sorted(d.items())
lst.append(('Итого', sum(d.values())))
wb.create_sheet('Result')
wb.active = wb['Result']
ws = wb.active
for i in range(1, len(lst) + 1):
    ws.cell(i, 1).value = lst[i-1][0]
    ws.cell(i, 2).value = lst[i-1][1]
wb.save('zp.xlsx')