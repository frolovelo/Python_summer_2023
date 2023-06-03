import openpyxl
name_file = input('Введите название файла с расширением: ')
wb = openpyxl.load_workbook(name_file)
ws = wb.active
lst = []
for i in range(1, ws.max_row + 1):
    lst.append(ws.cell(i, 2).value)
lst = sorted(lst)
wb.active = wb.create_sheet('Result')
ws = wb.active


def add_val(row, name, val):
    ws.cell(row, 1).value = name
    ws.cell(row, 2).value = val


add_val(1, 'Минимум', min(lst))
add_val(2, 'Максимум', max(lst))
add_val(3, 'Среднее арифметическое', round(sum(lst) / len(lst), 2))
add_val(4, 'Медиана', lst[len(lst) // 2] if len(lst) % 2 != 0 else round(sum(lst[len(lst) // 2 - 1:len(lst) // 2 + 1]) / 2, 2))

print('Проверяйте')
wb.save(name_file)
