'''openpyxl'''
import openpyxl

name_file = input('Введите название файла с расширением: ')
wb = openpyxl.load_workbook(name_file)
ws = wb.active
d = {}
for i in range(1, ws.max_row + 1):
    name = ws.cell(i, 1).value
    val = int(ws.cell(i, 2).value)
    d[name] = d.get(name, 0) + val
d['Итого'] = d.get('Итого', sum(d.values()))
wb.create_sheet('Result')
wb.active = wb['Result']
ws = wb.active
i = 1
for k, v in d.items():
    ws.cell(i, 1).value = k
    ws.cell(i, 2).value = v
    i += 1
print('Проверяйте')
wb.save(name_file)
